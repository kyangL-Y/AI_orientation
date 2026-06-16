#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
课程导入Excel模板生成工具
快速生成符合系统要求的Excel导入模板
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

def create_excel_template(output_file='课程导入模板.xlsx'):
    """创建课程导入Excel模板"""
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "课程数据"
    
    # 定义表头
    headers = [
        "主标题",
        "分类", 
        "具体分类",
        "课程名称",
        "知识点",
        "课程封面",
        "学习层级",
        "排序"
    ]
    
    # 设置表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # 设置列宽
    column_widths = {
        'A': 15,  # 主标题
        'B': 15,  # 分类
        'C': 15,  # 具体分类
        'D': 40,  # 课程名称
        'E': 50,  # 知识点
        'F': 30,  # 课程封面
        'G': 12,  # 学习层级
        'H': 8    # 排序
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 为学习层级列添加数据验证（下拉列表）
    from openpyxl.worksheet.datavalidation import DataValidation
    
    dv = DataValidation(type="list", formula1='"基础,进阶,高级"', allow_blank=True)
    dv.error = '请从下拉列表中选择：基础、进阶或高级'
    dv.errorTitle = '输入错误'
    dv.prompt = '请选择学习层级'
    dv.promptTitle = '学习层级'
    # 应用到G列（学习层级列），从第2行开始到第1000行
    dv.add(f"G2:G1000")
    ws.add_data_validation(dv)
    
    # 添加示例数据（第2-4行）
    example_data = [
        ["基础运营", "加入携程", "关于携程", "1.1.1携程集团（Trip.com Group）全面知识总结", "携程集团简介、业务模式、发展历程", "", "基础", 1],
        ["基础运营", "加入携程", "加盟须知", "1.2.1 服务缺陷操作手册", "服务标准、缺陷处理流程", "", "基础", 2],
        ["线上营销", "促销活动", "常规促销", "高级促销策略", "促销活动策划、执行技巧", "", "进阶", 10],
    ]
    
    for row_idx, row_data in enumerate(example_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_idx == 7:  # 学习层级列
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 8:  # 排序列
                cell.alignment = Alignment(horizontal="center")
    
    # 设置示例数据行的样式（浅灰色背景）
    example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for row in range(2, len(example_data) + 2):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = example_fill
    
    # 添加说明信息（在数据区域下方）
    note_row = len(example_data) + 3
    ws.merge_cells(f'A{note_row}:H{note_row + 5}')
    note_cell = ws.cell(row=note_row, column=1)
    note_cell.value = """填写说明：
1. 主标题、分类、课程名称为必填项
2. 学习层级可选：基础、进阶、高级（留空默认为"基础"）
3. 课程封面为图片URL，可为空
4. 排序值默认为0
5. 示例数据可以删除，从第2行开始填写您的数据
6. 学习层级列已设置下拉列表，可直接选择"""
    note_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    note_cell.font = Font(size=10, color="666666")
    note_cell.fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存文件
    wb.save(output_file)
    print(f"✅ Excel模板已生成：{output_file}")
    print(f"📝 文件位置：{os.path.abspath(output_file)}")
    print(f"💡 提示：模板已包含示例数据，可以直接修改或删除后填写您的数据")

if __name__ == "__main__":
    try:
        create_excel_template()
    except ImportError:
        print("❌ 错误：缺少 openpyxl 库")
        print("💡 请运行：pip install openpyxl")
    except Exception as e:
        print(f"❌ 生成模板时出错：{e}")

